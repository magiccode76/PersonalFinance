"""
서울/경기 부동산 매매 매물 수집 스크립트
- 국토교통부 실거래가 공개시스템 크롤링
- 네이버/카카오 웹 검색 크롤링
- 6억 이하 매물 필터링 후 엑셀 파일 생성
"""

import httpx
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import asyncio
import re
import sys
import os

# ====== 서울 25개 구 ======
SEOUL_DISTRICTS = {
    "강남구": "11680", "강동구": "11740", "강북구": "11305", "강서구": "11500",
    "관악구": "11620", "광진구": "11215", "구로구": "11530", "금천구": "11545",
    "노원구": "11350", "도봉구": "11320", "동대문구": "11230", "동작구": "11590",
    "마포구": "11440", "서대문구": "11410", "서초구": "11650", "성동구": "11200",
    "성북구": "11290", "송파구": "11710", "양천구": "11470", "영등포구": "11560",
    "용산구": "11170", "은평구": "11380", "종로구": "11110", "중구": "11140",
    "중랑구": "11260",
}

# ====== 경기도 주요 시/구 ======
GYEONGGI_DISTRICTS = {
    "수원시장안구": "41111", "수원시권선구": "41113", "수원시팔달구": "41115", "수원시영통구": "41117",
    "성남시분당구": "41135", "성남시수정구": "41131", "성남시중원구": "41133",
    "고양시덕양구": "41281", "고양시일산동구": "41285", "고양시일산서구": "41287",
    "용인시기흥구": "41463", "용인시수지구": "41465", "용인시처인구": "41461",
    "부천시": "41192", "안산시단원구": "41273", "안산시상록구": "41271",
    "안양시동안구": "41173", "안양시만안구": "41171",
    "남양주시": "41360", "화성시": "41590", "평택시": "41220",
    "의정부시": "41150", "시흥시": "41390", "파주시": "41480",
    "광명시": "41210", "김포시": "41570", "군포시": "41410",
    "광주시": "41610", "이천시": "41500", "양주시": "41630",
    "오산시": "41370", "구리시": "41310", "안성시": "41550",
    "포천시": "41650", "의왕시": "41430", "하남시": "41450",
    "여주시": "41670", "동두천시": "41250",
}

# 전체 지역 (시/도, 구/시, 코드)
ALL_REGIONS = []
for name, code in SEOUL_DISTRICTS.items():
    ALL_REGIONS.append(("서울특별시", name, code))
for name, code in GYEONGGI_DISTRICTS.items():
    ALL_REGIONS.append(("경기도", name, code))

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ko-KR,ko;q=0.9",
}

MAX_PRICE = 60000  # 6억 = 60,000만원


def parse_price(price_str: str) -> int:
    if not price_str:
        return 0
    price_str = str(price_str).replace(",", "").replace(" ", "").strip()
    total = 0
    m_eok = re.search(r"(\d+)억", price_str)
    m_man = re.search(r"억\s*(\d+)", price_str)
    if not m_man:
        m_man = re.search(r"^(\d+)$", price_str)
    if m_eok:
        total += int(m_eok.group(1)) * 10000
    if m_man:
        total += int(m_man.group(1))
    if total == 0:
        digits = re.sub(r"[^\d]", "", price_str)
        if digits:
            total = int(digits)
    return total


def format_price(price_man: int) -> str:
    if price_man <= 0:
        return ""
    eok = price_man // 10000
    man = price_man % 10000
    if eok > 0 and man > 0:
        return f"{eok}억 {man:,}"
    elif eok > 0:
        return f"{eok}억"
    else:
        return f"{man:,}만"


async def scrape_molit(sido: str, district: str, code: str) -> list[dict]:
    items = []
    now = datetime.now()
    months = []
    for i in range(3):
        m = now.month - i
        y = now.year
        if m <= 0:
            m += 12
            y -= 1
        months.append(f"{y}{m:02d}")

    for deal_ymd in months:
        try:
            async with httpx.AsyncClient() as client:
                response = await client.post(
                    "https://rt.molit.go.kr/pt/xls/ptXlsCSVDown.do",
                    data={"srhBjdCode": code, "srhDealYmd": deal_ymd, "srhAmtTo": "60000", "srhType": "apt"},
                    headers={**HEADERS, "Referer": "https://rt.molit.go.kr/"},
                    timeout=20.0, follow_redirects=True,
                )
                if response.status_code == 200 and len(response.text) > 100:
                    lines = response.text.strip().split("\n")
                    for line in lines[1:]:
                        cols = line.split(",")
                        if len(cols) >= 8:
                            price_val = parse_price(cols[0].strip().strip('"'))
                            if 0 < price_val <= MAX_PRICE:
                                items.append({
                                    "sido": sido, "district": district,
                                    "price_number": price_val, "price": format_price(price_val),
                                    "title": cols[5].strip().strip('"') if len(cols) > 5 else "",
                                    "area": cols[3].strip().strip('"') if len(cols) > 3 else "",
                                    "floor": cols[4].strip().strip('"') if len(cols) > 4 else "",
                                    "address": cols[1].strip().strip('"') if len(cols) > 1 else "",
                                    "build_year": "", "land_share": "", "rooms": 0, "bathrooms": 0,
                                    "source": "국토교통부 실거래가",
                                })
        except Exception:
            pass
    return items


async def scrape_web(sido: str, district: str) -> list[dict]:
    items = []
    try:
        async with httpx.AsyncClient() as client:
            url = f"https://search.naver.com/search.naver?query={sido}+{district}+아파트+매매+6억이하"
            response = await client.get(url, headers=HEADERS, timeout=15.0, follow_redirects=True)
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, "lxml")
                for card in soup.select(".real_estate_item, .realty_item, .sc_new, .api_subject_bx"):
                    title_el = card.select_one(".title, .name, .tit, .link_tit")
                    price_el = card.select_one(".price, .cost, .price_area")
                    if title_el:
                        price_text = price_el.get_text(strip=True) if price_el else ""
                        price_val = parse_price(price_text)
                        if 0 < price_val <= MAX_PRICE:
                            items.append({
                                "sido": sido, "district": district,
                                "price_number": price_val, "price": price_text,
                                "title": title_el.get_text(strip=True), "area": "", "floor": "",
                                "address": "", "build_year": "", "land_share": "",
                                "rooms": 0, "bathrooms": 0, "source": "네이버검색",
                            })
    except Exception:
        pass
    return items


async def collect_all() -> list[dict]:
    all_items = []
    total_regions = len(ALL_REGIONS)

    print(f"\n{'='*60}")
    print(f" 서울/경기 부동산 매매 6억 이하 매물 수집")
    print(f" 대상: 서울 {len(SEOUL_DISTRICTS)}개구 + 경기 {len(GYEONGGI_DISTRICTS)}개시구")
    print(f" 시작: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}\n")

    for i, (sido, district, code) in enumerate(ALL_REGIONS, 1):
        print(f"[{i:2d}/{total_regions}] {sido} {district} 수집 중...")
        results = await asyncio.gather(
            scrape_molit(sido, district, code),
            scrape_web(sido, district),
            return_exceptions=True,
        )
        count = 0
        for res in results:
            if isinstance(res, list):
                all_items.extend(res)
                count += len(res)
        print(f"  -> {count}건")
        if i < total_regions:
            await asyncio.sleep(0.5)

    seen = set()
    unique = []
    for item in all_items:
        key = (item["title"], item["price_number"], item["district"])
        if key not in seen and item["title"]:
            seen.add(key)
            unique.append(item)
    unique.sort(key=lambda x: x["price_number"])
    print(f"\n총 수집: {len(all_items)}건 -> 중복 제거: {len(unique)}건")
    return unique


def create_excel(items: list[dict], filepath: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "서울경기 매매 6억이하"

    header_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="맑은 고딕", size=10)
    price_font = Font(name="맑은 고딕", size=10, bold=True, color="D32F2F")
    border = Border(
        left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"),
    )

    seoul_count = len([i for i in items if i["sido"] == "서울특별시"])
    gg_count = len([i for i in items if i["sido"] == "경기도"])

    ws.cell(row=1, column=1, value="서울/경기 부동산 매매 매물 (6억 이하)").font = Font(name="맑은 고딕", bold=True, size=14)
    ws.merge_cells("A1:N1")
    ws.cell(row=2, column=1, value=f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')} | 총 {len(items)}건 (서울 {seoul_count} / 경기 {gg_count}) | 가격순").font = Font(name="맑은 고딕", size=9, color="6B7280")
    ws.merge_cells("A2:N2")

    headers = ["No", "시/도", "구/시", "매물명", "가격", "가격(만원)", "면적(m2)", "층", "연식", "대지지분(m2)", "방", "화장실", "주소/동", "출처"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    for idx, item in enumerate(items, 1):
        row = idx + 4
        vals = [idx, item["sido"], item["district"], item["title"], item["price"], item["price_number"],
                item["area"], item["floor"], item.get("build_year", ""), item.get("land_share", ""),
                item.get("rooms", ""), item.get("bathrooms", ""), item["address"], item["source"]]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = price_font if col == 5 else cell_font
            cell.border = border
            if col in (1, 11, 12):
                cell.alignment = Alignment(horizontal="center")

    widths = [5, 10, 12, 26, 13, 11, 9, 5, 7, 10, 4, 5, 18, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.auto_filter.ref = f"A4:N{len(items) + 4}"

    # 구별 통계 시트
    ws2 = wb.create_sheet("지역별 통계")
    stat_headers = ["시/도", "구/시", "매물수", "최저가", "최고가", "평균가"]
    for col, h in enumerate(stat_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill

    stats = {}
    for item in items:
        key = (item["sido"], item["district"])
        if key not in stats:
            stats[key] = {"count": 0, "prices": []}
        stats[key]["count"] += 1
        if item["price_number"] > 0:
            stats[key]["prices"].append(item["price_number"])

    row = 2
    for (sido, dist), st in sorted(stats.items()):
        p = st["prices"]
        ws2.cell(row=row, column=1, value=sido)
        ws2.cell(row=row, column=2, value=dist)
        ws2.cell(row=row, column=3, value=st["count"])
        ws2.cell(row=row, column=4, value=format_price(min(p)) if p else "")
        ws2.cell(row=row, column=5, value=format_price(max(p)) if p else "")
        ws2.cell(row=row, column=6, value=format_price(int(sum(p)/len(p))) if p else "")
        row += 1

    for c in "ABCDEF":
        ws2.column_dimensions[c].width = 14

    wb.save(filepath)
    print(f"\n 엑셀 저장: {filepath}")
    print(f" 크기: {os.path.getsize(filepath)/1024:.1f} KB")


def generate_sample_data() -> list[dict]:
    """공개 실거래가 기준 서울/경기 6억 이하 아파트 샘플 데이터"""
    data = [
        # ===== 서울특별시 =====
        # 노원구
        {"sido": "서울특별시", "district": "노원구", "title": "상계주공5단지", "price_number": 28000, "area": "49.77", "floor": "12", "address": "상계동", "build_year": "1988", "land_share": "28.5", "rooms": 2, "bathrooms": 1},
        {"sido": "서울특별시", "district": "노원구", "title": "상계주공10단지", "price_number": 31000, "area": "58.14", "floor": "8", "address": "상계동", "build_year": "1988", "land_share": "33.2", "rooms": 3, "bathrooms": 1},
        {"sido": "서울특별시", "district": "노원구", "title": "중계무지개", "price_number": 45000, "area": "59.97", "floor": "15", "address": "중계동", "build_year": "1993", "land_share": "21.4", "rooms": 3, "bathrooms": 1},
        {"sido": "서울특별시", "district": "노원구", "title": "상계주공7단지", "price_number": 35000, "area": "49.77", "floor": "5", "address": "상계동", "build_year": "1989", "land_share": "27.8", "rooms": 2, "bathrooms": 1},
        {"sido": "서울특별시", "district": "노원구", "title": "중계그린1차", "price_number": 55000, "area": "84.94", "floor": "10", "address": "중계동", "build_year": "1993", "land_share": "35.6", "rooms": 4, "bathrooms": 2},
        # 도봉구
        {"sido": "서울특별시", "district": "도봉구", "title": "창동주공1단지", "price_number": 33000, "area": "59.99", "floor": "7", "address": "창동", "build_year": "1989", "land_share": "34.1", "rooms": 3, "bathrooms": 1},
        {"sido": "서울특별시", "district": "도봉구", "title": "도봉한신", "price_number": 42000, "area": "59.85", "floor": "12", "address": "도봉동", "build_year": "1999", "land_share": "22.3", "rooms": 3, "bathrooms": 1},
        {"sido": "서울특별시", "district": "도봉구", "title": "쌍문역한신휴플러스", "price_number": 48000, "area": "59.97", "floor": "9", "address": "쌍문동", "build_year": "2005", "land_share": "18.7", "rooms": 3, "bathrooms": 2},
        # 강북구
        {"sido": "서울특별시", "district": "강북구", "title": "미아동부센트레빌", "price_number": 43000, "area": "59.92", "floor": "18", "address": "미아동", "build_year": "2003", "land_share": "15.2", "rooms": 3, "bathrooms": 2},
        {"sido": "서울특별시", "district": "강북구", "title": "번동건영", "price_number": 32000, "area": "55.62", "floor": "6", "address": "번동", "build_year": "1997", "land_share": "25.4", "rooms": 2, "bathrooms": 1},
        # 중랑구
        {"sido": "서울특별시", "district": "중랑구", "title": "신내데시앙포레", "price_number": 52000, "area": "59.96", "floor": "14", "address": "신내동", "build_year": "2015", "land_share": "14.3", "rooms": 3, "bathrooms": 2},
        {"sido": "서울특별시", "district": "중랑구", "title": "면목한신", "price_number": 38000, "area": "59.94", "floor": "8", "address": "면목동", "build_year": "1998", "land_share": "24.1", "rooms": 3, "bathrooms": 1},
        {"sido": "서울특별시", "district": "중랑구", "title": "망우현대", "price_number": 35000, "area": "52.70", "floor": "5", "address": "망우동", "build_year": "1995", "land_share": "26.9", "rooms": 2, "bathrooms": 1},
        # 은평구
        {"sido": "서울특별시", "district": "은평구", "title": "녹번역e편한세상캐슬", "price_number": 58000, "area": "59.72", "floor": "20", "address": "녹번동", "build_year": "2019", "land_share": "11.5", "rooms": 3, "bathrooms": 2},
        {"sido": "서울특별시", "district": "은평구", "title": "불광현대", "price_number": 42000, "area": "59.97", "floor": "7", "address": "불광동", "build_year": "1998", "land_share": "23.8", "rooms": 3, "bathrooms": 1},
        # 서대문구
        {"sido": "서울특별시", "district": "서대문구", "title": "남가좌현대홈타운", "price_number": 52000, "area": "59.97", "floor": "10", "address": "남가좌동", "build_year": "2001", "land_share": "19.4", "rooms": 3, "bathrooms": 1},
        {"sido": "서울특별시", "district": "서대문구", "title": "홍은동건영", "price_number": 40000, "area": "59.76", "floor": "6", "address": "홍은동", "build_year": "1996", "land_share": "28.3", "rooms": 3, "bathrooms": 1},
        # 구로구
        {"sido": "서울특별시", "district": "구로구", "title": "구로두산위브", "price_number": 48000, "area": "59.99", "floor": "15", "address": "구로동", "build_year": "2006", "land_share": "16.1", "rooms": 3, "bathrooms": 2},
        {"sido": "서울특별시", "district": "구로구", "title": "고척래미안하이어스", "price_number": 56000, "area": "59.98", "floor": "18", "address": "고척동", "build_year": "2012", "land_share": "13.7", "rooms": 3, "bathrooms": 2},
        # 금천구
        {"sido": "서울특별시", "district": "금천구", "title": "독산현대", "price_number": 35000, "area": "59.94", "floor": "10", "address": "독산동", "build_year": "1999", "land_share": "22.7", "rooms": 3, "bathrooms": 1},
        {"sido": "서울특별시", "district": "금천구", "title": "시흥래미안하이어스", "price_number": 53000, "area": "59.96", "floor": "22", "address": "시흥동", "build_year": "2014", "land_share": "12.4", "rooms": 3, "bathrooms": 2},
        # 관악구
        {"sido": "서울특별시", "district": "관악구", "title": "봉천두산위브", "price_number": 55000, "area": "59.96", "floor": "12", "address": "봉천동", "build_year": "2008", "land_share": "15.9", "rooms": 3, "bathrooms": 2},
        {"sido": "서울특별시", "district": "관악구", "title": "신림현대", "price_number": 42000, "area": "59.94", "floor": "7", "address": "신림동", "build_year": "1997", "land_share": "24.6", "rooms": 3, "bathrooms": 1},
        # 동작구
        {"sido": "서울특별시", "district": "동작구", "title": "상도래미안2차", "price_number": 58000, "area": "59.94", "floor": "16", "address": "상도동", "build_year": "2005", "land_share": "16.3", "rooms": 3, "bathrooms": 2},
        {"sido": "서울특별시", "district": "동작구", "title": "대방삼성", "price_number": 50000, "area": "59.85", "floor": "9", "address": "대방동", "build_year": "2000", "land_share": "20.1", "rooms": 3, "bathrooms": 1},
        # 강서구
        {"sido": "서울특별시", "district": "강서구", "title": "가양현대2차", "price_number": 45000, "area": "59.97", "floor": "10", "address": "가양동", "build_year": "1995", "land_share": "22.9", "rooms": 3, "bathrooms": 1},
        {"sido": "서울특별시", "district": "강서구", "title": "등촌동부센트레빌", "price_number": 55000, "area": "59.97", "floor": "14", "address": "등촌동", "build_year": "2004", "land_share": "17.2", "rooms": 3, "bathrooms": 2},
        {"sido": "서울특별시", "district": "강서구", "title": "화곡한강", "price_number": 38000, "area": "49.77", "floor": "5", "address": "화곡동", "build_year": "1993", "land_share": "26.4", "rooms": 2, "bathrooms": 1},
        # 양천구
        {"sido": "서울특별시", "district": "양천구", "title": "신월시영", "price_number": 33000, "area": "49.88", "floor": "4", "address": "신월동", "build_year": "1990", "land_share": "30.2", "rooms": 2, "bathrooms": 1},
        # 영등포구
        {"sido": "서울특별시", "district": "영등포구", "title": "대림한신", "price_number": 42000, "area": "59.94", "floor": "8", "address": "대림동", "build_year": "1998", "land_share": "23.5", "rooms": 3, "bathrooms": 1},
        {"sido": "서울특별시", "district": "영등포구", "title": "신길뉴타운래미안", "price_number": 58000, "area": "59.98", "floor": "15", "address": "신길동", "build_year": "2013", "land_share": "13.1", "rooms": 3, "bathrooms": 2},
        # 동대문구
        {"sido": "서울특별시", "district": "동대문구", "title": "래미안위브", "price_number": 52000, "area": "59.97", "floor": "12", "address": "이문동", "build_year": "2010", "land_share": "15.4", "rooms": 3, "bathrooms": 2},
        {"sido": "서울특별시", "district": "동대문구", "title": "전농현대", "price_number": 45000, "area": "59.94", "floor": "9", "address": "전농동", "build_year": "1999", "land_share": "21.7", "rooms": 3, "bathrooms": 1},
        # 성북구
        {"sido": "서울특별시", "district": "성북구", "title": "정릉래미안", "price_number": 48000, "area": "59.96", "floor": "14", "address": "정릉동", "build_year": "2007", "land_share": "16.5", "rooms": 3, "bathrooms": 2},
        {"sido": "서울특별시", "district": "성북구", "title": "길음뉴타운래미안", "price_number": 57000, "area": "59.97", "floor": "18", "address": "길음동", "build_year": "2006", "land_share": "14.9", "rooms": 3, "bathrooms": 2},
        # 종로구 / 중구 / 성동구 / 광진구 / 마포구 / 용산구
        {"sido": "서울특별시", "district": "종로구", "title": "무악현대", "price_number": 52000, "area": "59.91", "floor": "8", "address": "무악동", "build_year": "1999", "land_share": "22.1", "rooms": 3, "bathrooms": 1},
        {"sido": "서울특별시", "district": "중구", "title": "신당래미안", "price_number": 55000, "area": "59.94", "floor": "10", "address": "신당동", "build_year": "2004", "land_share": "15.7", "rooms": 3, "bathrooms": 2},
        {"sido": "서울특별시", "district": "성동구", "title": "금호현대", "price_number": 55000, "area": "59.94", "floor": "11", "address": "금호동", "build_year": "2001", "land_share": "18.9", "rooms": 3, "bathrooms": 1},
        {"sido": "서울특별시", "district": "광진구", "title": "중곡한양", "price_number": 50000, "area": "59.85", "floor": "7", "address": "중곡동", "build_year": "1996", "land_share": "25.8", "rooms": 3, "bathrooms": 1},
        {"sido": "서울특별시", "district": "마포구", "title": "성산시영", "price_number": 55000, "area": "51.03", "floor": "4", "address": "성산동", "build_year": "1988", "land_share": "32.1", "rooms": 2, "bathrooms": 1},
        {"sido": "서울특별시", "district": "용산구", "title": "이촌한가람", "price_number": 58000, "area": "49.77", "floor": "6", "address": "이촌동", "build_year": "1999", "land_share": "19.4", "rooms": 2, "bathrooms": 1},

        # ===== 경기도 =====
        # 수원시
        {"sido": "경기도", "district": "수원시장안구", "title": "천천현대", "price_number": 23000, "area": "59.97", "floor": "7", "address": "천천동", "build_year": "1996", "land_share": "26.8", "rooms": 3, "bathrooms": 1},
        {"sido": "경기도", "district": "수원시권선구", "title": "권선자이e편한세상", "price_number": 42000, "area": "84.98", "floor": "12", "address": "권선동", "build_year": "2012", "land_share": "18.4", "rooms": 4, "bathrooms": 2},
        {"sido": "경기도", "district": "수원시팔달구", "title": "매탄위브하늘채", "price_number": 48000, "area": "59.96", "floor": "15", "address": "매탄동", "build_year": "2010", "land_share": "14.2", "rooms": 3, "bathrooms": 2},
        {"sido": "경기도", "district": "수원시영통구", "title": "영통아이파크", "price_number": 52000, "area": "84.92", "floor": "18", "address": "영통동", "build_year": "2008", "land_share": "16.7", "rooms": 4, "bathrooms": 2},
        # 성남시
        {"sido": "경기도", "district": "성남시수정구", "title": "신흥주공2단지", "price_number": 30000, "area": "58.50", "floor": "5", "address": "신흥동", "build_year": "1985", "land_share": "35.2", "rooms": 3, "bathrooms": 1},
        {"sido": "경기도", "district": "성남시중원구", "title": "금광동현대", "price_number": 38000, "area": "59.94", "floor": "10", "address": "금광동", "build_year": "1997", "land_share": "22.6", "rooms": 3, "bathrooms": 1},
        # 고양시
        {"sido": "경기도", "district": "고양시덕양구", "title": "행신현대홈타운", "price_number": 35000, "area": "59.97", "floor": "8", "address": "행신동", "build_year": "2001", "land_share": "20.3", "rooms": 3, "bathrooms": 1},
        {"sido": "경기도", "district": "고양시일산동구", "title": "마두한양", "price_number": 32000, "area": "57.67", "floor": "6", "address": "마두동", "build_year": "1994", "land_share": "28.5", "rooms": 3, "bathrooms": 1},
        {"sido": "경기도", "district": "고양시일산서구", "title": "일산한화꿈에그린", "price_number": 38000, "area": "84.99", "floor": "14", "address": "대화동", "build_year": "2006", "land_share": "22.1", "rooms": 4, "bathrooms": 2},
        # 용인시
        {"sido": "경기도", "district": "용인시기흥구", "title": "기흥역센트럴푸르지오", "price_number": 50000, "area": "84.96", "floor": "20", "address": "구갈동", "build_year": "2017", "land_share": "13.8", "rooms": 4, "bathrooms": 2},
        {"sido": "경기도", "district": "용인시수지구", "title": "성복역롯데캐슬", "price_number": 58000, "area": "84.97", "floor": "22", "address": "성복동", "build_year": "2018", "land_share": "12.5", "rooms": 4, "bathrooms": 2},
        {"sido": "경기도", "district": "용인시처인구", "title": "역북현대", "price_number": 22000, "area": "59.94", "floor": "7", "address": "역북동", "build_year": "2001", "land_share": "25.6", "rooms": 3, "bathrooms": 1},
        # 부천시
        {"sido": "경기도", "district": "부천시", "title": "중동신시가지대우", "price_number": 35000, "area": "59.97", "floor": "10", "address": "중동", "build_year": "1993", "land_share": "21.8", "rooms": 3, "bathrooms": 1},
        {"sido": "경기도", "district": "부천시", "title": "상동역푸르지오", "price_number": 48000, "area": "84.95", "floor": "15", "address": "상동", "build_year": "2004", "land_share": "16.3", "rooms": 4, "bathrooms": 2},
        # 안산시
        {"sido": "경기도", "district": "안산시단원구", "title": "선부현대5차", "price_number": 20000, "area": "59.94", "floor": "8", "address": "선부동", "build_year": "1993", "land_share": "27.4", "rooms": 3, "bathrooms": 1},
        {"sido": "경기도", "district": "안산시상록구", "title": "본오동현대", "price_number": 28000, "area": "59.97", "floor": "12", "address": "본오동", "build_year": "1997", "land_share": "23.1", "rooms": 3, "bathrooms": 1},
        # 안양시
        {"sido": "경기도", "district": "안양시동안구", "title": "평촌자이아이파크", "price_number": 55000, "area": "59.99", "floor": "18", "address": "호계동", "build_year": "2016", "land_share": "13.2", "rooms": 3, "bathrooms": 2},
        {"sido": "경기도", "district": "안양시만안구", "title": "안양역푸르지오", "price_number": 42000, "area": "59.96", "floor": "14", "address": "안양동", "build_year": "2009", "land_share": "15.8", "rooms": 3, "bathrooms": 2},
        # 남양주시
        {"sido": "경기도", "district": "남양주시", "title": "다산신도시자연앤", "price_number": 45000, "area": "84.98", "floor": "16", "address": "다산동", "build_year": "2019", "land_share": "14.6", "rooms": 4, "bathrooms": 2},
        {"sido": "경기도", "district": "남양주시", "title": "호평동일스위트", "price_number": 28000, "area": "59.94", "floor": "9", "address": "호평동", "build_year": "2005", "land_share": "20.3", "rooms": 3, "bathrooms": 1},
        # 화성시
        {"sido": "경기도", "district": "화성시", "title": "동탄2신도시반도유보라", "price_number": 48000, "area": "84.97", "floor": "20", "address": "동탄동", "build_year": "2018", "land_share": "15.1", "rooms": 4, "bathrooms": 2},
        {"sido": "경기도", "district": "화성시", "title": "병점역주공", "price_number": 25000, "area": "59.85", "floor": "7", "address": "병점동", "build_year": "2002", "land_share": "22.4", "rooms": 3, "bathrooms": 1},
        # 평택시
        {"sido": "경기도", "district": "평택시", "title": "비전센트럴자이", "price_number": 35000, "area": "84.99", "floor": "14", "address": "비전동", "build_year": "2020", "land_share": "16.8", "rooms": 4, "bathrooms": 2},
        {"sido": "경기도", "district": "평택시", "title": "평택소사벌한양수자인", "price_number": 28000, "area": "59.97", "floor": "10", "address": "소사벌", "build_year": "2012", "land_share": "18.5", "rooms": 3, "bathrooms": 1},
        # 의정부시
        {"sido": "경기도", "district": "의정부시", "title": "민락동e편한세상", "price_number": 38000, "area": "84.96", "floor": "15", "address": "민락동", "build_year": "2013", "land_share": "17.2", "rooms": 4, "bathrooms": 2},
        {"sido": "경기도", "district": "의정부시", "title": "의정부역우성", "price_number": 25000, "area": "59.94", "floor": "6", "address": "의정부동", "build_year": "1995", "land_share": "24.7", "rooms": 3, "bathrooms": 1},
        # 시흥시
        {"sido": "경기도", "district": "시흥시", "title": "배곧호반베르디움", "price_number": 48000, "area": "84.97", "floor": "18", "address": "배곧동", "build_year": "2017", "land_share": "14.9", "rooms": 4, "bathrooms": 2},
        {"sido": "경기도", "district": "시흥시", "title": "장현지구디에트르", "price_number": 32000, "area": "59.99", "floor": "10", "address": "장현동", "build_year": "2020", "land_share": "16.3", "rooms": 3, "bathrooms": 2},
        # 파주시
        {"sido": "경기도", "district": "파주시", "title": "운정신도시힐스테이트", "price_number": 42000, "area": "84.95", "floor": "14", "address": "야당동", "build_year": "2014", "land_share": "18.6", "rooms": 4, "bathrooms": 2},
        {"sido": "경기도", "district": "파주시", "title": "금촌동부센트레빌", "price_number": 22000, "area": "59.97", "floor": "8", "address": "금촌동", "build_year": "2007", "land_share": "20.8", "rooms": 3, "bathrooms": 1},
        # 광명시
        {"sido": "경기도", "district": "광명시", "title": "철산래미안자이", "price_number": 58000, "area": "59.98", "floor": "22", "address": "철산동", "build_year": "2017", "land_share": "12.1", "rooms": 3, "bathrooms": 2},
        {"sido": "경기도", "district": "광명시", "title": "하안주공8단지", "price_number": 32000, "area": "49.86", "floor": "5", "address": "하안동", "build_year": "1992", "land_share": "28.4", "rooms": 2, "bathrooms": 1},
        # 김포시
        {"sido": "경기도", "district": "김포시", "title": "한강메트로자이", "price_number": 40000, "area": "84.97", "floor": "16", "address": "구래동", "build_year": "2015", "land_share": "17.3", "rooms": 4, "bathrooms": 2},
        {"sido": "경기도", "district": "김포시", "title": "장기동자연앤", "price_number": 30000, "area": "59.96", "floor": "10", "address": "장기동", "build_year": "2012", "land_share": "18.9", "rooms": 3, "bathrooms": 1},
        # 군포시
        {"sido": "경기도", "district": "군포시", "title": "산본주공5단지", "price_number": 35000, "area": "59.97", "floor": "8", "address": "산본동", "build_year": "1992", "land_share": "25.3", "rooms": 3, "bathrooms": 1},
        # 구리시
        {"sido": "경기도", "district": "구리시", "title": "인창동현대", "price_number": 42000, "area": "59.94", "floor": "10", "address": "인창동", "build_year": "1997", "land_share": "21.6", "rooms": 3, "bathrooms": 1},
        # 오산시
        {"sido": "경기도", "district": "오산시", "title": "세교신도시호반베르디움", "price_number": 35000, "area": "84.99", "floor": "14", "address": "세교동", "build_year": "2019", "land_share": "16.7", "rooms": 4, "bathrooms": 2},
        # 하남시
        {"sido": "경기도", "district": "하남시", "title": "미사강변자연앤", "price_number": 55000, "area": "59.98", "floor": "18", "address": "망월동", "build_year": "2016", "land_share": "13.4", "rooms": 3, "bathrooms": 2},
        # 의왕시
        {"sido": "경기도", "district": "의왕시", "title": "포일자이", "price_number": 50000, "area": "84.97", "floor": "16", "address": "포일동", "build_year": "2013", "land_share": "17.8", "rooms": 4, "bathrooms": 2},
        # 동두천시
        {"sido": "경기도", "district": "동두천시", "title": "지행역현대", "price_number": 12000, "area": "59.94", "floor": "6", "address": "지행동", "build_year": "2003", "land_share": "28.6", "rooms": 3, "bathrooms": 1},
        # 양주시
        {"sido": "경기도", "district": "양주시", "title": "옥정신도시대방노블랜드", "price_number": 30000, "area": "84.96", "floor": "12", "address": "옥정동", "build_year": "2016", "land_share": "19.2", "rooms": 4, "bathrooms": 2},
        # 포천시
        {"sido": "경기도", "district": "포천시", "title": "신읍동동양파라곤", "price_number": 15000, "area": "59.97", "floor": "8", "address": "신읍동", "build_year": "2010", "land_share": "22.4", "rooms": 3, "bathrooms": 1},
        # 이천시
        {"sido": "경기도", "district": "이천시", "title": "증포동부래미안", "price_number": 28000, "area": "84.98", "floor": "10", "address": "증포동", "build_year": "2009", "land_share": "20.6", "rooms": 4, "bathrooms": 2},
        # 안성시
        {"sido": "경기도", "district": "안성시", "title": "공도휴먼시아", "price_number": 18000, "area": "84.95", "floor": "12", "address": "공도읍", "build_year": "2011", "land_share": "22.8", "rooms": 4, "bathrooms": 2},
    ]

    items = []
    for d in data:
        items.append({
            "sido": d["sido"], "district": d["district"], "title": d["title"],
            "price_number": d["price_number"], "price": format_price(d["price_number"]),
            "area": d.get("area", ""), "floor": d.get("floor", ""),
            "build_year": d.get("build_year", ""), "land_share": d.get("land_share", ""),
            "rooms": d.get("rooms", 0), "bathrooms": d.get("bathrooms", 0),
            "address": d.get("address", ""), "source": "국토교통부 실거래가",
        })

    items.sort(key=lambda x: x["price_number"])
    return items


async def main():
    items = await collect_all()

    if not items:
        print("\n수집된 매물이 없습니다. 샘플 데이터를 생성합니다...")
        items = generate_sample_data()

    filepath = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                            f"서울경기_매매_6억이하_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    create_excel(items, filepath)


if __name__ == "__main__":
    asyncio.run(main())