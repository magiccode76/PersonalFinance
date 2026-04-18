"""
엑셀 파일의 매물 데이터를 MongoDB에 저장하는 스크립트
사용법: python3 scripts/import_excel_to_db.py [엑셀파일경로]
"""

import sys
import os
from datetime import datetime
from pymongo import MongoClient
from openpyxl import load_workbook

MONGODB_URL = os.environ.get(
    "MONGODB_URL",
    "mongodb://pfuser:pfpass123@localhost:27017/personalfinance?authSource=personalfinance"
)
MONGODB_DB = os.environ.get("MONGODB_DB", "personalfinance")


def import_excel(filepath: str):
    print(f"\n=== 엑셀 -> MongoDB 임포트 ===")
    print(f"파일: {filepath}")

    # 엑셀 읽기
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active

    # 헤더 확인 (4행)
    headers = [cell.value for cell in ws[4]]
    print(f"헤더: {headers}")

    # 데이터 읽기 (5행부터)
    items = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row[0]:
            continue
        # 헤더: No, 구, 매물명, 가격, 가격(만원), 면적, 층, 연식, 대지지분, 방, 화장실, 주소/동, 출처
        items.append({
            "title": str(row[2] or ""),
            "property_type": "아파트",
            "trade_type": "매매",
            "price": str(row[3] or ""),
            "price_number": int(row[4] or 0),
            "area": float(row[5]) if row[5] else 0.0,
            "floor": str(row[6] or ""),
            "build_year": str(row[7] or "") if len(row) > 7 else "",
            "land_share": str(row[8] or "") if len(row) > 8 else "",
            "rooms": int(row[9] or 0) if len(row) > 9 else 0,
            "bathrooms": int(row[10] or 0) if len(row) > 10 else 0,
            "address": str(row[11] or "") if len(row) > 11 else "",
            "region": f"서울특별시 {row[1]}",
            "description": "",
            "source": str(row[12] or "국토교통부 실거래가") if len(row) > 12 else "국토교통부 실거래가",
            "source_url": "https://rt.molit.go.kr",
            "image_url": "",
            "created_at": datetime.now(),
            "updated_at": datetime.now(),
        })

    print(f"읽은 데이터: {len(items)}건")

    if not items:
        print("데이터가 없습니다.")
        return

    # MongoDB 연결 및 저장
    client = MongoClient(MONGODB_URL)
    db = client[MONGODB_DB]

    # 기존 국토교통부 데이터 삭제 (중복 방지)
    deleted = db.properties.delete_many({"source": {"$regex": "국토교통부|공공데이터"}})
    print(f"기존 데이터 삭제: {deleted.deleted_count}건")

    # 새 데이터 삽입
    result = db.properties.insert_many(items)
    print(f"새 데이터 저장: {len(result.inserted_ids)}건")

    # 확인
    total = db.properties.count_documents({})
    print(f"전체 DB 매물 수: {total}건")

    client.close()
    print("완료!\n")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        filepath = sys.argv[1]
    else:
        # 가장 최근 엑셀 파일 찾기
        project_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        xlsx_files = sorted(
            [f for f in os.listdir(project_dir) if f.endswith(".xlsx") and "서울" in f],
            reverse=True
        )
        if xlsx_files:
            filepath = os.path.join(project_dir, xlsx_files[0])
        else:
            print("엑셀 파일을 찾을 수 없습니다. 경로를 인자로 전달해주세요.")
            sys.exit(1)

    import_excel(filepath)
